from flask import Flask, render_template, request, redirect, url_for, flash, send_from_directory,session
import pandas as pd
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime
from openpyxl import Workbook, load_workbook
import requests
import logging
import json
import os
import re
import openpyxl
from flask_sqlalchemy import SQLAlchemy
from volcenginesdkarkruntime import Ark


# 配置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


app = Flask(__name__,
            static_folder='assets',
            static_url_path='/assets')
app.secret_key = 'your_secret_key'

path1 = None
audio_file_path=None
text_to_pass=""
# 自定义静态文件路由
@app.route('/vendors/<path:filename>')
def custom_static(filename):
    return send_from_directory('vendors', filename)

# 定义 Excel 文件路径
EXCEL_FILE = "styles.xlsx"

# 使用绝对路径
USERS_FILE = r"D:\pyprojects\public\users.xlsx"


# 检查文件是否存在，如果不存在则创建文件并添加初始用户
if not os.path.exists(USERS_FILE):
    logger.error("Error: The specified file does not exist. Creating file...")
    # 定义初始用户数据
    data = {
        'username': ['admin'],
        'password': [generate_password_hash('1234')],  # 生成密码哈希
        'role': ['student']
    }
    df = pd.DataFrame(data)
    df.to_excel(USERS_FILE, index=False)
    logger.info("File created and user added.")
print(f"The path to the users file is: {USERS_FILE}")

#------------------------------------- Teacher 界面函数 -------------------------------------
# Teacher:初始化 Excel 文件
def init_excel():
    if not os.path.exists(EXCEL_FILE):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Styles"
        # 设置表头，新增 AudioPath 和 FileName
        sheet.append(["ID", "LanguageStyle", "ClassStyle", "Timestamp", "AudioPath", "FileName","OutLine"])
        workbook.save(EXCEL_FILE)

# Teacher:更新插入数据函数，增加 AudioPath 和 FileName 字段
def insert_style_to_excel(language_style, class_style, audio_path, file_name,outline):
    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active
    # 检查 ID 列中的实际有效行数，忽略空行
    next_id = len([row for row in sheet.iter_rows(min_col=1, max_col=1, values_only=True) if row[0] is not None]) + 1
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    # 插入数据：ID, 两个值, 时间戳, 音频路径和文件名
    sheet.append([next_id, language_style, class_style, timestamp, audio_path, file_name,outline])
    workbook.save(EXCEL_FILE)

# Teacher:从表格中拿取数据并上传
@app.route('/get_style/<int:style_id>', methods=['GET'])
def get_style(style_id):
    # 打开 Excel 文件，获取对应的行
    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=True):  # 跳过表头，从第二行开始
        if row[0] == style_id:
            # 返回 JSON 数据
            return {
                "id": row[0],
                "language_style": row[1],
                "class_style": row[2],
                "timestamp": row[3],
                "audio_path":row[4],
                "file_name":row[5],
                "outline":row[6],

            }
    return {"error": "Style not found"}, 404

# Teacher:获取所有存储的 styles
def get_all_styles_from_excel():
    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))[1:]  # 跳过表头
    workbook.close()
    return rows

#------------------------------------- Upload 界面函数 -------------------------------------

# Upload:音频转文本函数
@app.route('/transcribe_audio', methods=['POST'])
def get_audio_transcription():
    global audio_file_path
    global text_to_pass
    url = "https://api.siliconflow.cn/v1/audio/transcriptions"
    model_name = "FunAudioLLM/SenseVoiceSmall"
    # 从前端请求中获取上传的音频文件
    audio_file = request.files.get('audio_file')
    if audio_file:
        # 保存音频文件到临时位置（这里假设保存到当前目录下的temp文件夹，可根据实际调整）
        if not os.path.exists('temp'):
            os.makedirs('temp')
        audio_file_path = os.path.join('temp', audio_file.filename)
        audio_file.save(audio_file_path)
        print(audio_file_path)
        # 构建要发送的文件数据和其他表单数据
        files = {
            'file': (audio_file_path, open(audio_file_path, 'rb')),
            'model': (None, model_name)
        }
        headers = {
            "Authorization": "Bearer sk-puyhjrhxfvuxewqvnjpxuhrrzzflbtfcojlvcaruvnqxehyi"
        }

        try:
            response = requests.post(url, files=files, headers=headers)
            response.raise_for_status()
            data = json.loads(response.text)
            # 提取正文内容
            text = data.get("text", "")
            print("here")
            print(text)
            text_to_pass=text
            return text
        except requests.exceptions.RequestException as e:
            print(f"请求出现错误: {e}")
            return "请求出现错误，请稍后再试"
    return "没有接收到有效的音频文件"

'''@app.route('/teacher_next', methods=['POST', 'GET'])
def teacher_next():
    # 从 POST 请求中获取 text_to_pass 参数
    text_to_pass = request.form.get('text_to_pass', '').strip()
    print(f"接收到的文本内容: {text_to_pass}")

    if not text_to_pass:
        return "错误：未接收到文本内容", 400

    # 将文本内容保存到 session
    session['text_to_pass'] = text_to_pass
    print("保存到 session 中的内容:", session['text_to_pass'])

    # 重定向到 teacher 页面
    return redirect(url_for('teacher'))'''

#------------------------------------- Home 界面函数 -------------------------------------
# Home:匹配账号函数
def user_exists(username):
    try:
        df = pd.read_excel(USERS_FILE)
        return not df[df['username'] == username].empty
    except Exception as e:
        logger.error(f"无法读取Excel文件：{e}")
        return False

# Home:匹配密码函数
def check_password(username, password):
    try:
        df = pd.read_excel(USERS_FILE)
        user = df[df['username'] == username]
        if not user.empty:
            # 强制转换密码列为字符串类型
            password_hash = str(user['password'].iloc[0])
            return check_password_hash(password_hash, password)
    except Exception as e:
        logger.error(f"无法验证密码：{e}")
        return False

# Home:匹配角色函数
def get_user_role(username):
    try:
        df = pd.read_excel(USERS_FILE)
        user = df[df['username'] == username]
        if not user.empty:
            return str(user['role'].iloc[0])  # 确保角色也是字符串类型
    except Exception as e:
        logger.error(f"无法获取用户角色：{e}")
        return None

# Home:注册后添加信息
def add_user_to_excel(username, password, role):
    try:
        print(f"Adding user: {username}, {role}")
        df = pd.read_excel(USERS_FILE)  # 尝试读取文件
        print(f"Current DataFrame: \n{df}")
        new_user = pd.DataFrame({'username': [username], 'password': [password], 'role': [role]})
        df = pd.concat([df, new_user], ignore_index=True)
        df.to_excel(USERS_FILE, index=False)  # 尝试写入文件
        print(f"Updated DataFrame: \n{df}")
        logger.info(f"用户 {username} 已成功添加到Excel文件中。")
        return True
    except Exception as e:
        logger.error(f"无法将用户 {username} 添加到Excel文件中：{e}")
        print(f"Error: {e}")
        return False

#------------------------------------- Teacher 路由 -------------------------------------
# Teacher:页面路由

@app.route('/teacher', methods=['GET', 'POST'])
def teacher():
    if request.method == 'POST':
        # 处理表单提交的数据
        # 例如，获取上传的文件或其他表单数据
        pass

    # 从 generate 模块中获取两个值
    language_style, class_style = trans()  # 修改 generate 返回两个值
    audio_path = audio_file_path.replace('temp', 'assets/video')
    outline = out()
    file_name = os.path.basename(audio_file_path)
    insert_style_to_excel(language_style, class_style, audio_path,file_name, outline)
    all_styles = get_all_styles_from_excel()  # 获取所有历史记录
    return render_template('teacher.html', language_style=language_style, class_style=class_style,
                            all_styles=all_styles, audio_path=audio_path, outline=outline,file_name=file_name)

'''@app.route('/teacher', methods=['GET', 'POST'])
def teacher():
    if request.method == 'POST':
        # 处理表单提交的数据
        # 例如，获取上传的文件或其他表单数据
        pass
    print("nowhere")
    print(text_to_pass)
    audio_path = audio_file_path
    all_styles = get_all_styles_from_excel()
    return render_template('teacher.html',all_styles=all_styles,text_test=text_to_pass,audio_path=audio_path)'''

def generate():
    # 获取 text_to_pass 的值
    textget = text_to_pass
    if textget is None:
        textget = ''  # 或者其他默认值

    # 设置 Ark 客户端
    client = Ark(api_key=os.environ.get("ARK_API_KEY"))
    client = Ark(
        base_url="https://ark.cn-beijing.volces.com/api/v3",
    )
    input = '请你分析一下这段上课的音频，总结教师的教学风格和语言风格。' + textget +'请严格按照以下格式输出：整体语言风格（10个字及以下）： 。课堂风格分析及建议（200字及以下 ）： 。'
    completion = client.chat.completions.create(
        model="ep-20241207202047-vv45w",
        messages=[
            {"role": "system", "content": "你是一个专业的教授，现在需要来评判不同教师上课的教学风格。"},
            {"role": "user", "content": input},
        ],
    )
    return completion.choices[0].message.content

def outline():
    # 设置 Ark 客户端
    textget = text_to_pass
    client = Ark(api_key=os.environ.get("ARK_API_KEY"))
    client = Ark(
        base_url="https://ark.cn-beijing.volces.com/api/v3",
    )
    input = '请你分析一下这段上课的音频文本，总结教师的课程大纲' + textget +'请严格按照以下格式输出，字数在200字左右，不要有序号：课程大纲： 。'
    completion = client.chat.completions.create(
        model="ep-20241207202047-vv45w",
        messages=[
            {"role": "system", "content": "你是一个专业的教授，现在需要来记录不同老师的课程大纲。"},
            {"role": "user", "content": input},
        ],
    )
    return completion.choices[0].message.content

def parse_generated_content(content):
    pattern = r"整体语言风格：\s*(.*?)。\s*课堂风格分析及建议：\s*(.*)"
    match = re.search(pattern, content)
    if match:
        language_style = match.group(1)
        class_style = match.group(2)
        print(language_style)
        return language_style, class_style
    else:
        return None, None

def parse_generated_outline(content):
    pattern = r"课程大纲：\s*(.*)"
    match = re.search(pattern, content)
    if match:
        dagang = match.group(1)
        print(dagang)
        return dagang
    else:
        return None, None

def trans():
    language_style, class_style = parse_generated_content(generate())
    return language_style,class_style

def out():
    outlinefinal = parse_generated_outline(outline())
    return outlinefinal
#------------------------------------- Student 路由 -------------------------------------

def get_access_token():
    """
    使用 API Key，Secret Key 获取access_token，替换下列示例中的应用API Key、应用Secret Key
    """
    url = "https://aip.baidubce.com/oauth/2.0/token?grant_type=client_credentials&client_id=GnArphNvN1w7vv2gQmR1Zzbn&client_secret=Y4cKfqiLqQyufWnZBta8Hj3AipkyMgux"
    payload = json.dumps("")
    headers = {
        'Content-Type': 'application/json',
        'Accept': 'application/json'
    }
    response = requests.request("POST", url, headers=headers, data=payload)
    return response.json().get("access_token")

def generate_student_notes(text):
    if text is None or not isinstance(text, str):
        print("传入的text变量无效，请检查其获取逻辑。")
        raise ValueError("无效的文本输入")

    access_token = get_access_token()
    note_generation_url = f"https://aip.baidubce.com/rpc/2.0/ai_custom/v1/wenxinworkshop/chat/qianfan-agent-speed-8k?access_token={access_token}"
    payload = json.dumps({
        "messages": [
            {
                "role": "user",
                "content": "请根据以下提供的课堂文本内容生成学生笔记，笔记要求输出三个方面，首先是要点，以关键词、简短标题等形式，第二个是笔记，记录重点内容，借助符号、缩写等形式，言简意赅。第三个是总结，写下感受、思考和体会。呈现课堂文本内容如下：" + text
            }
        ]
    })
    headers = {
        'Content-Type': 'application/json'
    }

    try:
        response = requests.request("POST", note_generation_url, headers=headers, data=payload)
        response.raise_for_status()
        data = response.json()
        if 'result' in data:
            student_notes = data['result']
            # 先按照换行符进行初步分割，因为每个部分之间大概率是换行隔开的
            parts_by_line = student_notes.splitlines()
            points = ""
            note = ""
            summary = ""
            flag_points = False
            flag_note = False
            flag_summary = False
            for line in parts_by_line:
                if "要点" in line:
                    flag_points = True
                    continue
                elif "笔记" in line:
                    flag_note = True
                    flag_points = False
                    continue
                elif "总结" in line:
                    flag_summary = True
                    flag_note = False
                    continue
                if flag_points:
                    points += line + "\n"
                elif flag_note:
                    note += line + "\n"
                elif flag_summary:
                    summary += line + "\n"
            # 去除末尾多余的换行符
            points = points.rstrip()
            note = note.rstrip()
            summary = summary.rstrip()
            return points, note, summary
        else:
            print("接口返回数据格式不符合预期，无法提取学生笔记内容。")
            return "", "", ""
    except requests.RequestException as e:
        if isinstance(e, requests.ConnectionError):
            print("网络连接出现问题，请检查网络设置。")
        elif isinstance(e, requests.Timeout):
            print("请求超时，请稍后再试或检查网络稳定性。")
        elif isinstance(e, requests.HTTPError):
            print(f"HTTP请求返回错误状态码，详细信息: {e.response.text}")
        else:
            print(f"其他请求相关错误: {e}")
        return "", "", ""


def generate_graph_student_notes(text):
    """
    根据输入文本生成适合用于知识图谱的学生笔记格式，包含学科、实体、关系等信息
    """
    if text is None or not isinstance(text, str):
        print("传入的text变量无效，请检查其获取逻辑。")
        raise ValueError("无效的文本输入")

    access_token = get_access_token()
    graph_note_generation_url = f"https://aip.baidubce.com/rpc/2.0/ai_custom/v1/wenxinworkshop/chat/qianfan-agent-speed-8k?access_token={access_token}"
    # 这里提示语调整为要求大模型按照期望格式生成笔记，示例中以语文举例，可根据实际传入学科信息调整
    payload = json.dumps({
        "messages": [
            {
                "role": "user",
                "content": "请根据以下提供的课堂文本内容生成适合用于知识图谱的学生笔记，格式要求如下：\n"
                           "{\n"
                           "    \"subject\": \"语文\",\n"
                           "    \"entities\": [具体的实体内容，多个实体用逗号隔开],\n"
                           "    \"relations\": [[实体1, 关系描述, 实体2], [实体3, 关系描述, 实体4]]\n"
                           "}\n"
                           "呈现课堂文本内容如下：" + text
            }
        ]
    })
    headers = {
        'Content-Type': 'application/json'
    }

    try:
        response = requests.request("POST", graph_note_generation_url, headers=headers, data=payload)
        response.raise_for_status()
        data = response.json()
        if 'result' in data:
            graph_student_notes = data['result']
            print(graph_student_notes)
            print(type(graph_student_notes))
            return graph_student_notes
        else:
            print("接口返回数据格式不符合预期，无法提取用于知识图谱的学生笔记内容。")
            return {}
    except requests.RequestException as e:
        if isinstance(e, requests.ConnectionError):
            print("网络连接出现问题，请检查网络设置。")
        elif isinstance(e, requests.Timeout):
            print("请求超时，请稍后再试或检查网络稳定性。")
        elif isinstance(e, requests.HTTPError):
            print(f"HTTP请求返回错误状态码，详细信息: {e.response.text}")
        else:
            print(f"其他请求相关错误: {e}")
        return {}

# Student:页面路由
@app.route('/student', methods=['GET', 'POST'])
def student():
    excel_file_path = os.path.join(os.getcwd(), "student_notes.xlsx")    # 定义Excel文件路径
    if request.method == 'POST':
        # 获取表单数据，并去除前后空格（优化获取后的数据处理）
        points = request.form.get('points', '').strip()
        note = request.form.get('note', '').strip()
        summary = request.form.get('summary', '').strip()
        subject = request.form.get('subject', '')
        note_type = request.form.get('note_type', '')

        # 数据验证（示例简单验证，可根据实际需求完善）
        if not all([points, note, summary, subject, note_type]):
            return render_template('index.html', error_message="请填写完整的笔记信息", audio_path=audio_file_path)

        # 判断Excel文件是否存在，不存在则创建并添加表头
        if not os.path.exists(excel_file_path):
            wb = Workbook()
            ws = wb.active
            ws.append(["要点", "笔记", "总结", "学科", "笔记类型","路径"])
        else:
            # 使用load_workbook函数加载已有文件，而非通过Workbook实例调用
            wb = load_workbook(excel_file_path)
            ws = wb.active

        # 将数据添加到Excel表格中
        ws.append([points, note, summary, subject, note_type,audio_file_path])

        try:
            # 保存Excel文件
            wb.save(excel_file_path)
            return render_template('student.html', success_message="笔记保存成功", audio_path=audio_file_path)
        except:
            return render_template('student.html', error_message="笔记保存失败，请稍后重试", audio_path=audio_file_path)

    # 获取音频转录文本及后续处理逻辑等
    audio_transcription = text_to_pass
    if audio_transcription:
        # 简单验证，去除可能的恶意脚本标签等（示例，可根据实际完善）
        audio_transcription = re.sub('<[^<]+?>', '', audio_transcription)
        # 根据转录文本生成学生笔记
        audio_path = audio_file_path.replace('temp', 'assets/video')
        student_notes = generate_student_notes(audio_transcription)
        if isinstance(student_notes, tuple):
            points, note, summary = student_notes
            return render_template('student.html', points=points, note=note, summary= summary, audio_path=audio_path)
        else:
            return render_template('student.html', error_message="生成学生笔记的返回格式不符合预期", audio_path=audio_file_path)
    else:
        return render_template('student.html', error_message="未能成功获取音频转录文本", audio_path=audio_file_path)


@app.route('/show_matched_content')
def show_matched_content():
    # 从查询参数中获取匹配要点内容，如果没有传入则默认为空字符串
    target_points = "语文"
    excel_file_path = "student_notes.xlsx"
    matched_notes = []

    if not os.path.exists(excel_file_path):
        return render_template(
            'content.html',
            error_message="Excel文件不存在，无法进行匹配操作",
            matched_notes=[]
        )

    try:
        wb = openpyxl.load_workbook(excel_file_path)
        ws = wb.active

        # 遍历表格中的每一行（跳过表头行，从第二行开始），检查要点列是否包含要匹配的要点内容
        for row in ws.iter_rows(min_row=2, values_only=True):
            if target_points in str(row[3]):
                matched_notes.append({
                    "points": row[0],
                    "note": row[1],
                    "summary": row[2],
                    "subject": row[3],
                    "category": row[4],
                    "audio_path": "assets/video/example.mp3"  # 替换为实际音频路径逻辑
                })

        return render_template(
            'content.html',
            matched_notes=matched_notes  # 将匹配的笔记传递到模板
        )

    except Exception as e:
        return render_template(
            'content.html',
            error_message=f"匹配笔记过程中出现错误: {str(e)}",
            matched_notes=[]
        )

@app.route('/graph')
def graph():
    audio_transcription = text_to_pass
    graph_student_notes = generate_student_notes.generate_graph_student_notes(audio_transcription)
    return render_template('nicegraph.html',graph_student_notes = generate_student_notes)


#------------------------------------- Upload 路由 -------------------------------------
# Upload:老师上传路由
@app.route('/upload_teacher', methods=['GET'])
def upload_teacher():
    return render_template('upload_teacher.html')

# Upload:学生上传路由
@app.route('/upload_student', methods=['GET'])
def upload_student():
        return render_template('upload_student.html')

#------------------------------------- Home 路由 -------------------------------------
# Home:主页路由
@app.route('/')
def home():
    return render_template('index.html')

# Home:从表中匹配账号密码信息 登陆路由
@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        role = request.form.get('role')  # 从表单中获取角色

        logger.info(f"Attempting to login with username: {username}")

        # 检查用户是否存在以及密码是否正确
        if not user_exists(username) or not check_password(username, password):
            error = '用户名或密码错误'
            logger.error(f"Login failed for user: {username}")
        else:
            # 获取用户角色
            user_role = get_user_role(username)
            if user_role is None or user_role != role:
                error = '角色不匹配'
                logger.error(f"Role mismatch for user: {username}")
            else:
                logger.info(f"Login successful for user: {username}")
                if user_role == "student":
                    return redirect(url_for('student_home'))
                elif user_role == "teacher":
                    return redirect(url_for('teacher_home'))

    return render_template('login.html', error=error)

# Home:生成并存储注册的密码 注册路由
@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        confirmPassword = request.form.get('confirmPassword')
        role = request.form.get('role')
        print(username)

        # 密码不一致
        if password != confirmPassword:
            flash('密码和确认密码不一致', 'error')
            return render_template('register.html')

        # 用户名已存在
        if user_exists(username):
            flash('用户名已存在', 'error')
            return render_template('register.html')

        # 生成哈希密码并添加用户
        hashed_password = generate_password_hash(password)
        if add_user_to_excel(username, hashed_password, role):
            flash('注册成功，请登录', 'success')
            return redirect(url_for('login'))
        else:
            flash('注册失败，请稍后再试', 'error')
            return render_template('register.html')

    return render_template('register.html')

# Home:学生预览路由
@app.route('/student_home',endpoint='student_home')
def student_home():
    return render_template('student_home.html')

# Home:教师预览路由
@app.route('/teacher_home')
def teacher_home():
    return render_template('teacher_home.html')

# Home:上传音频路由
@app.route('/audio')
def audio():
    return render_template('audio.html')


if __name__ == '__main__':
    app.run(debug=True)

